#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
note_font = Font(italic=True, color="7F6000", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ========== 구매담당자 리스트 ==========
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "Sheet1"

headers1 = ["연번", "소속", "부서", "성명", "휴대폰", "이메일", "분야", "세부품목", "관심기업(번호)", "관심기업(기업명)"]
notes1   = ["필수\nA1, A2...", "필수\n기관명 전체", "필수\n부서명", "필수\n성명", "선택", "선택",
            "필수\n관심품목 대분류", "필수\n관심품목 세부", "선택\n(미사용)", "핵심\n기업명 쉼표구분\n(파일B열과 정확히 일치)"]

sample_buyers = [
    ["A1", "한국환경공단 수도권동부환경본부", "환경서비스처 환경안전경영부", "홍길동", "010-1234-5678", "hong@keco.or.kr",
     "수질(펌프)", "수중펌프 및 교반기", "", "(주)워터테크, 한강수처리"],
    ["A2", "한국환경공단 부산울산경남환경본부", "환경서비스처 유해대기관리부", "김철수", "010-2345-6789", "kim@keco.or.kr",
     "측정분석기(대기)", "대기오염측정시스템", "", "대기환경(주)"],
    ["A3", "한국환경공단", "경영지원처 총무부", "이영희", "010-3456-7890", "lee@keco.or.kr",
     "폐기물, 탈취", "폐기물처리, 탈취기", "", "(주)클린환경, 에코솔루션, 그린테크"],
    ["A4", "한국수자원공사", "수자원관리부", "박민준", "010-4567-8901", "park@kwater.or.kr",
     "수질(필터)", "수처리용 필터, 정수막", "", "(주)워터필터, 정수기술"],
    ["A5", "서울물재생시설공단", "시설관리부", "최지원", "010-5678-9012", "choi@seoul.go.kr",
     "수처리", "하수처리 설비", "", "(주)하수처리기술, 물환경"],
    ["A6", "경기도시공사", "환경사업부", "정수현", "010-6789-0123", "jung@gico.or.kr",
     "재활용", "재생플라스틱 원료", "", ""],
    ["A7", "SH엔솔", "기술연구소", "강동원", "010-7890-1234", "kang@shenol.com",
     "에너지", "태양광 발전 시스템", "", "(주)솔라에너지"],
]

ws1.row_dimensions[1].height = 35
ws1.row_dimensions[2].height = 50

# 헤더
for c, h in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# 설명 행
for c, n in enumerate(notes1, 1):
    cell = ws1.cell(row=2, column=c, value=n)
    cell.fill = note_fill
    cell.font = note_font
    cell.alignment = center
    cell.border = border

# 샘플 데이터
for r, row_data in enumerate(sample_buyers, 3):
    ws1.row_dimensions[r].height = 20
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.alignment = left
        cell.border = border

col_widths1 = [10, 28, 25, 10, 16, 24, 18, 28, 14, 32]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws1.freeze_panes = "A3"
wb1.save("구매담당자_리스트_예시.xlsx")
print("[OK] 구매담당자_리스트_예시.xlsx 생성 완료")


# ========== 참가기업 리스트 ==========
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "설문 답변(1-50)"

headers2 = [
    "연번", "회사명", "세부품목", "대분류",
    "5월 21일 상담회 신청", "5월 22일 상담회 신청",
    "관심 바이어\n(환경공단)", "관심 바이어\n(수자원공사)",
    "관심 바이어\n(엔지니어링·발전사)", "관심 바이어\n(환경산업&탄소중립)",
    "관심 바이어\n(경기도)", "관심 바이어\n(자원순환·재활용)",
    "관심 바이어\n(서울물재생시설공단)", "우선 상담 희망"
]
notes2 = [
    "필수\n1,2,3...", "필수\n바이어J열과\n정확히 일치", "필수\n주요품목 서술", "필수\n제조/도소매 등",
    "참가시 해당\n상담회명 입력\n(비워두면 미참가)", "참가시 해당\n상담회명 입력\n(비워두면 미참가)",
    "환경공단\n바이어번호\n쉼표구분", "수자원공사\n바이어번호\n쉼표구분",
    "엔지니어링발전사\n바이어번호\n쉼표구분", "환경산업탄소\n바이어번호\n쉼표구분",
    "경기도\n바이어번호\n쉼표구분", "자원순환\n바이어번호\n쉼표구분",
    "서울물산업\n바이어번호\n쉼표구분", "선택\n(미사용)"
]

sample_companies = [
    [1, "(주)워터테크", "수중펌프, 교반기, 수처리 설비", "제조",
     "한국환경공단 구매상담회", "환경산업&탄소중립 공공 구매상담회",
     "1, 2, 3", "", "", "10, 11", "", "", "", "1, 2, 10"],
    [2, "한강수처리", "수질분석기, 수처리용 필터", "제조",
     "한국환경공단 구매상담회, 한국수자원공사 구매상담회", "",
     "1, 4, 5", "4", "", "", "", "", "", "1, 4"],
    [3, "대기환경(주)", "대기오염측정시스템, 환경측정기기", "제조",
     "한국환경공단 구매상담회", "경기도 환경산업 공공 구매상담회",
     "2", "", "", "", "20, 21", "", "", "2, 20"],
    [4, "(주)클린환경", "폐기물처리 설비, 탈취기", "제조,도소매",
     "한국환경공단 구매상담회, 엔지니어링·발전사 내수 구매상담회", "자원순환·재활용 구매상담회",
     "3", "", "30, 31", "", "", "40", "", "3, 30, 40"],
    [5, "에코솔루션", "탈취기, 악취방지시설", "제조",
     "한국환경공단 구매상담회", "",
     "3", "", "", "", "", "", "", "3"],
    [6, "(주)워터필터", "수처리용 필터, 정수막", "제조",
     "한국수자원공사 구매상담회", "서울물재생시설공단 물산업 구매상담회",
     "", "4", "", "", "", "", "5", "4, 5"],
    [7, "(주)솔라에너지", "태양광 발전 시스템, 에너지저장장치", "제조",
     "한국환경공단 구매상담회", "",
     "7", "", "", "", "", "", "", "7"],
]

ws2.row_dimensions[1].height = 45
ws2.row_dimensions[2].height = 55

for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for c, n in enumerate(notes2, 1):
    cell = ws2.cell(row=2, column=c, value=n)
    cell.fill = note_fill
    cell.font = note_font
    cell.alignment = center
    cell.border = border

for r, row_data in enumerate(sample_companies, 3):
    ws2.row_dimensions[r].height = 20
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = left
        cell.border = border

col_widths2 = [8, 18, 28, 12, 22, 22, 12, 12, 14, 14, 10, 12, 16, 14]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws2.freeze_panes = "A3"
wb2.save("참가기업_리스트_예시.xlsx")
print("[OK] 참가기업_리스트_예시.xlsx 생성 완료")
