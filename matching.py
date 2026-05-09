#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
from typing import Dict, List, Set, Tuple
from collections import defaultdict

# ==================== 데이터 로딩 ====================

# 5/21 상담회 참가 기관 목록
ORGS_MAY_21 = ['한국환경공단', '한국수자원공사', 'SH엔솔', '한국기술']

def determine_date(org: str) -> str:
    """소속 기관명으로 참가 날짜 결정"""
    for org_name in ORGS_MAY_21:
        if org_name in org:
            return '2026-05-21'
    return '2026-05-22'


def load_buyers(filepath):
    """
    바이어 리스트 로드 (날짜 정보, 관심 기업 포함)
    Returns: {buyer_id: {id, org, dept, name, interests, interest_tokens, date, interested_companies}}

    2026 컬럼 구조:
    A: 연번(buyer_id), B: 소속, C: 부서, D: 성명, E: 휴대폰, F: 이메일,
    G: 분야, H: 세부품목, I: 관심기업(번호), J: 관심기업(기업명)
    날짜: 소속 기관명으로 판별
    관심기업: 기업명(J열)으로 매칭
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Sheet1']

    buyers = {}
    warnings = []

    buyers_without_interested_companies = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
        buyer_id_cell = row[0].value          # A: 연번
        org = row[1].value if len(row) > 1 else ""   # B: 소속
        dept = row[2].value if len(row) > 2 else ""  # C: 부서
        name = row[3].value if len(row) > 3 else ""  # D: 성명
        # E: 휴대폰, F: 이메일 (미사용)
        field = row[6].value if len(row) > 6 else ""       # G: 분야
        detail = row[7].value if len(row) > 7 else ""      # H: 세부품목
        # I: 관심기업(번호) (미사용 - 번호 체계가 달라 이름으로 매칭)
        interested_companies_text = row[9].value if len(row) > 9 else ""  # J: 관심기업(기업명)

        if buyer_id_cell is None:
            break

        buyer_id = str(buyer_id_cell).strip()
        org = str(org).strip() if org else ""
        dept = str(dept).strip() if dept else ""
        name = str(name).strip() if name else ""
        field = str(field).strip() if field else ""
        detail = str(detail).strip() if detail else ""
        interested_companies_text = str(interested_companies_text).strip() if interested_companies_text else ""

        # 관심품목: 분야 + 세부품목 합산
        interests_text = ', '.join(filter(None, [field, detail]))

        # 날짜: 소속 기관명으로 결정
        date_str = determine_date(org)

        # 관심 기업명 추출 (쉼표, 개행으로 분할)
        interested_companies = set()
        if interested_companies_text:
            company_names = re.split(r'[,\n]', interested_companies_text)
            for cname in company_names:
                cname = cname.strip()
                if cname:
                    interested_companies.add(cname)
        else:
            buyers_without_interested_companies.append({
                'buyer_id': buyer_id,
                'name': name,
                'org': org
            })

        interest_tokens = tokenize(interests_text)

        buyers[buyer_id] = {
            'id': buyer_id,
            'org': org,
            'dept': dept,
            'name': name,
            'interests': interests_text,
            'interest_tokens': interest_tokens,
            'date': date_str,
            'interested_companies': interested_companies,
            'has_interested_companies': len(interested_companies) > 0
        }

    if buyers_without_interested_companies:
        for item in buyers_without_interested_companies:
            warnings.append(f"바이어 {item['buyer_id']} ({item['org']} {item['name']}): 관심 기업 없음")

    return buyers, warnings


def load_companies(filepath, buyers):
    """
    회사 리스트 로드
    Returns: {company_name: {seq, name, products, category, wished_buyer_ids, product_tokens, dates}}

    2026 컬럼 구조:
    A: 연번, B: 회사명, C: 세부품목, D: 대분류,
    E: 5/21 참가 여부(텍스트), F: 5/22 참가 여부(텍스트),
    G~M: 기관별 관심 구매담당자 번호, N: 우선 상담 희망 번호
    날짜: E/F열 유무로 결정
    관심바이어: G~M열 합산 후 숫자→"A{번호}" 변환
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb['설문 답변(1-50)']

    companies = {}
    warnings = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
        seq = row[0].value           # A: 연번
        name = row[1].value          # B: 회사명
        products = row[2].value      # C: 세부품목
        category = row[3].value      # D: 대분류
        attend_521 = row[4].value    # E: 5/21 참가 여부
        attend_522 = row[5].value    # F: 5/22 참가 여부
        # G(6)~M(12): 기관별 관심 구매담당자, N(13): 우선 상담 희망

        if seq is None:
            break

        seq_str = str(seq).strip()
        name = str(name).strip() if name else ""
        products = str(products).strip() if products else ""
        category = str(category).strip() if category else ""

        # 날짜 결정: E/F열 텍스트 유무로 판단
        company_dates = set()
        if attend_521:
            company_dates.add('2026-05-21')
        if attend_522:
            company_dates.add('2026-05-22')

        # 관심 바이어 수집: G~M 7개 컬럼 합산 (숫자 → "A{번호}")
        wished_buyer_ids = set()
        for col_idx in range(6, 13):  # G(6)~M(12)
            if col_idx >= len(row):
                continue
            val = row[col_idx].value
            if val is None:
                continue
            for part in re.split(r'[,\n]', str(val)):
                part = part.strip()
                if part.isdigit():
                    wished_buyer_ids.add(f'A{part}')

        product_tokens = tokenize(products)

        companies[name] = {
            'seq': seq_str,
            'name': name,
            'products': products,
            'category': category,
            'wished_buyer_ids': wished_buyer_ids,
            'product_tokens': product_tokens,
            'dates': company_dates,
            'category_tokens': {category} if category else set()
        }

    return companies, warnings


# ==================== 토크나이징 ====================

def tokenize(text):
    """
    텍스트를 토큰으로 분할
    - 쉼표, 슬래시, 개행, 괄호로 분할
    - 각 토큰 strip
    - 2글자 이상 토큰만 유지
    Returns: set of tokens
    """
    if not text:
        return set()

    # 괄호 내용은 별도 처리
    text = re.sub(r'\([^)]*\)', ' ', text)
    text = re.sub(r'\[[^\]]*\]', ' ', text)

    # 쉼표, 슬래시, 개행으로 분할
    tokens = re.split(r'[,/\n]', text)

    result = set()
    for token in tokens:
        token = token.strip()
        if len(token) >= 2:
            result.add(token)

    return result


# ==================== 매칭 판정 ====================

def buyer_wishes_company(buyer_tokens: Set[str], company_tokens: Set[str]) -> bool:
    """
    바이어의 관심품목 토큰이 기업의 주요품목에 포함되는지 판정
    - buyer_tokens 중 하나 이상이 company의 product 텍스트에 부분문자열로 포함되면 True
    """
    if not buyer_tokens or not company_tokens:
        return False

    for buyer_token in buyer_tokens:
        for company_token in company_tokens:
            if buyer_token in company_token or company_token in buyer_token:
                return True

    return False


# ==================== 풀 분류 ====================

def classify_pairs(buyers: Dict, companies: Dict):
    """
    (바이어, 기업) 쌍을 5개 풀로 분류 (같은 날짜에만 매칭)

    Pool1: 상호 관심 - 바이어 관심 기업 AND 기업의 관심 구매담당자 일치
    Pool2: 바이어 우선 - 바이어 관심 기업 (기업이 선택 안 해도 됨)
    Pool3: 기업 우선 - 기업의 관심 구매담당자 (바이어가 선택 안 해도 됨)
    Pool4: 세부품목 - 바이어 관심품목 AND 기업 주요품목 일치
    Pool5: 대분류 - 바이어 관심품목 AND 기업 품목분류 일치
    """
    pool1 = []  # 상호 관심
    pool2 = []  # 바이어 우선
    pool3 = []  # 기업 우선
    pool4 = []  # 세부품목
    pool5 = []  # 대분류

    for company_name, company in companies.items():
        for buyer_id, buyer in buyers.items():
            # 날짜 체크: 바이어와 기업이 같은 날짜에 있어야 함
            buyer_date = buyer.get('date', '')
            company_dates = company.get('dates', set())

            if buyer_date not in company_dates:
                continue  # 다른 날짜면 매칭 대상에서 제외

            # Pool 분류 로직 (2026: 기업명으로 매칭)
            buyer_interested_company = company['name'] in buyer['interested_companies']
            company_interested_buyer = buyer_id in company['wished_buyer_ids']
            buyer_matches_products = buyer_wishes_company(buyer['interest_tokens'], company['product_tokens'])
            buyer_matches_category = buyer_wishes_company(buyer['interest_tokens'], company['category_tokens'])

            # Pool1: 상호 관심 (바이어가 관심 기업을 지정한 경우만)
            if buyer['has_interested_companies'] and buyer_interested_company and company_interested_buyer:
                pool1.append((buyer_id, company_name))
            # Pool2: 바이어 우선 (바이어가 관심 기업을 지정한 경우만)
            elif buyer['has_interested_companies'] and buyer_interested_company and not company_interested_buyer:
                pool2.append((buyer_id, company_name))
            # Pool3: 기업 우선 (바이어가 관심 기업을 지정하지 않아도 됨)
            elif company_interested_buyer and not buyer_interested_company:
                pool3.append((buyer_id, company_name))
            # Pool4: 세부품목 일치
            elif buyer_matches_products:
                pool4.append((buyer_id, company_name))
            # Pool5: 대분류 일치
            elif buyer_matches_category:
                pool5.append((buyer_id, company_name))

    return pool1, pool2, pool3, pool4, pool5


# ==================== 그리디 매칭 ====================

def greedy_matching(buyers: Dict, companies: Dict, pool1, pool2, pool3, pool4, pool5):
    """
    그리디 라운드 배정 알고리즘 (5개 풀)
    - 각 바이어는 최대 4개 슬롯
    - 각 라운드에서 처리 제약:
      [C1] 바이어 슬롯 소진 시 배정 불가
      [C2] 동일 (바이어, 기업) 쌍 중복 배정 불가
      [C3] 동일 바이어가 같은 라운드에 2회 배정 불가
      [C4] 동일 기업이 같은 라운드에 2회 배정 불가
    """
    slots = {buyer_id: 4 for buyer_id in buyers.keys()}
    matched = set()  # {(buyer_id, company_name)}
    buyer_in_round = {1: set(), 2: set(), 3: set(), 4: set()}
    company_in_round = {1: set(), 2: set(), 3: set(), 4: set()}

    matches_by_round = {1: [], 2: [], 3: [], 4: []}
    match_type = {}  # {(buyer_id, company_name): 'pool1'|'pool2'|...|'pool5'}

    # 각 풀에서 처리 (순서: pool1 → pool2 → pool3 → pool4 → pool5)
    for pool_name, pool in [('pool1', pool1), ('pool2', pool2), ('pool3', pool3), ('pool4', pool4), ('pool5', pool5)]:
        changed = True
        while changed:
            changed = False
            for r in [1, 2, 3, 4]:
                for (buyer_id, company_name) in pool:
                    # [C3] 바이어가 이미 이 라운드에 배정되었나?
                    if buyer_id in buyer_in_round[r]:
                        continue

                    # [C4] 기업이 이미 이 라운드에 배정되었나?
                    if company_name in company_in_round[r]:
                        continue

                    # [C2] 이 쌍이 이미 배정되었나?
                    if (buyer_id, company_name) in matched:
                        continue

                    # [C1] 바이어 슬롯이 남았나?
                    if slots[buyer_id] == 0:
                        continue

                    # 배정 성공
                    matches_by_round[r].append((buyer_id, company_name))
                    matched.add((buyer_id, company_name))
                    buyer_in_round[r].add(buyer_id)
                    company_in_round[r].add(company_name)
                    slots[buyer_id] -= 1
                    match_type[(buyer_id, company_name)] = pool_name
                    changed = True

    return matches_by_round, matched, match_type, slots


# ==================== 기관별 라운드 배정 (매칭유형 포함) 시트 생성 ====================

def create_organization_with_type_sheet(wb, buyers, companies, matches_by_round, matched, match_type):
    """
    바이어 소속 기관별로 분리하여 라운드별 매칭 기업과 매칭유형을 색상으로 표시
    구조: 참가날짜 | 기관 | 부서 | 이름 | 관심품목 | 라운드1(색상) | 라운드2(색상) | 라운드3(색상) | 라운드4(색상)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 첫 번째 시트 생성
    ws = wb.create_sheet("기관별 라운드 배정 (매칭유형)", 0)

    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    pool1_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pool2_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    pool3_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # 바이어를 날짜-기관-부서별로 그룹화
    buyers_by_date_org = {}
    for buyer_id, buyer in buyers.items():
        date = buyer.get('date', '')
        org = buyer['org']

        if date not in buyers_by_date_org:
            buyers_by_date_org[date] = {}
        if org not in buyers_by_date_org[date]:
            buyers_by_date_org[date][org] = []

        buyers_by_date_org[date][org].append((buyer_id, buyer))

    # 각 기관 내에서 바이어를 부서 및 이름으로 정렬
    for date in buyers_by_date_org:
        for org in buyers_by_date_org[date]:
            buyers_by_date_org[date][org].sort(key=lambda x: (x[1]['dept'], x[1]['name']))

    # 헤더 작성
    headers = ["참가날짜", "소속 기관", "부서", "성명", "관심품목", "라운드1", "라운드2", "라운드3", "라운드4"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 데이터 입력 (날짜순, 기관순)
    row = 2
    for date in sorted(buyers_by_date_org.keys()):
        for org in sorted(buyers_by_date_org[date].keys()):
            for buyer_id, buyer in buyers_by_date_org[date][org]:
                # 라운드별 매칭 기업과 유형 수집
                round_data = {1: [], 2: [], 3: [], 4: []}
                for r in [1, 2, 3, 4]:
                    for match_buyer_id, match_company_name in matches_by_round[r]:
                        if match_buyer_id == buyer_id:
                            m_type = match_type[(match_buyer_id, match_company_name)]
                            round_data[r].append((match_company_name, m_type))

                # 행 데이터 입력
                ws.cell(row=row, column=1, value=buyer.get('date', ''))
                ws.cell(row=row, column=2, value=buyer['org'])
                ws.cell(row=row, column=3, value=buyer['dept'])
                ws.cell(row=row, column=4, value=buyer['name'])
                ws.cell(row=row, column=5, value=buyer['interests'])

                # 각 라운드의 매칭 기업들을 색상으로 표시
                for r in [1, 2, 3, 4]:
                    companies_text = "\n".join([f"{comp}" for comp, _ in round_data[r]])
                    cell = ws.cell(row=row, column=5+r, value=companies_text)

                    # 매칭 유형에 따라 색상 지정 (첫 번째 매칭 유형으로 색상 결정)
                    if round_data[r]:
                        first_type = round_data[r][0][1]
                        if first_type == 'pool1':
                            cell.fill = pool1_fill
                        elif first_type == 'pool2':
                            cell.fill = pool2_fill
                        elif first_type == 'pool3':
                            cell.fill = pool3_fill

                # 스타일 적용
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if col <= 5:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

                row += 1

    # 열 너비 설정
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    for col in range(6, 10):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # 기본 행 높이
    ws.row_dimensions[1].height = 25


# ==================== 기관별 라운드 배정 시트 생성 ====================

def create_organization_sheet(wb, buyers, companies, matches_by_round, matched, match_type):
    """
    바이어 소속 기관별로 분리하여 라운드별 매칭 기업을 표시
    구조: 기관 | 부서 | 이름 | 관심품목 | 라운드1 | 라운드2 | 라운드3 | 라운드4
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 첫 번째 시트 생성
    ws = wb.create_sheet("기관별 라운드 배정", 0)

    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    org_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    org_header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # 바이어를 날짜-기관-부서별로 그룹화
    buyers_by_date_org = {}
    for buyer_id, buyer in buyers.items():
        date = buyer.get('date', '')
        org = buyer['org']

        if date not in buyers_by_date_org:
            buyers_by_date_org[date] = {}
        if org not in buyers_by_date_org[date]:
            buyers_by_date_org[date][org] = []

        buyers_by_date_org[date][org].append((buyer_id, buyer))

    # 각 기관 내에서 바이어를 부서 및 이름으로 정렬
    for date in buyers_by_date_org:
        for org in buyers_by_date_org[date]:
            buyers_by_date_org[date][org].sort(key=lambda x: (x[1]['dept'], x[1]['name']))

    # 헤더 작성 (참가날짜를 맨 앞으로)
    headers = ["참가날짜", "소속 기관", "부서", "성명", "관심품목", "라운드1", "라운드2", "라운드3", "라운드4"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 데이터 입력 (날짜순, 기관순)
    row = 2
    for date in sorted(buyers_by_date_org.keys()):
        for org in sorted(buyers_by_date_org[date].keys()):
            org_start_row = row

            for buyer_id, buyer in buyers_by_date_org[date][org]:
                # 라운드별 매칭 기업 수집
                round_companies = {1: [], 2: [], 3: [], 4: []}
                for r in [1, 2, 3, 4]:
                    for match_buyer_id, match_company_name in matches_by_round[r]:
                        if match_buyer_id == buyer_id:
                            round_companies[r].append(match_company_name)

                # 행 데이터 입력 (참가날짜를 맨 앞으로)
                ws.cell(row=row, column=1, value=buyer.get('date', ''))
                ws.cell(row=row, column=2, value=buyer['org'])
                ws.cell(row=row, column=3, value=buyer['dept'])
                ws.cell(row=row, column=4, value=buyer['name'])
                ws.cell(row=row, column=5, value=buyer['interests'])

                # 각 라운드의 매칭 기업들을 줄바꿈으로 표시
                for r in [1, 2, 3, 4]:
                    companies_text = "\n".join(round_companies[r])
                    ws.cell(row=row, column=5+r, value=companies_text)

                # 스타일 적용
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if col <= 5:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

                row += 1

        # 기관 구분선 (선택사항 - 색상으로 구분)
        if row < ws.max_row + 1:
            pass

    # 열 너비 설정
    ws.column_dimensions['A'].width = 12  # 참가날짜
    ws.column_dimensions['B'].width = 18  # 소속 기관
    ws.column_dimensions['C'].width = 15  # 부서
    ws.column_dimensions['D'].width = 12  # 성명
    ws.column_dimensions['E'].width = 25  # 관심품목
    for col in range(6, 10):  # 라운드 1-4
        ws.column_dimensions[get_column_letter(col)].width = 25

    # 기본 행 높이
    ws.row_dimensions[1].height = 25
    for r in range(2, row):
        ws.row_dimensions[r].height = None  # 자동 높이 (wrap_text 때문에)


# ==================== 결과 출력 ====================

def write_results(output_filepath, buyers, companies, matches_by_round, matched, match_type, slots, load_warnings):
    """
    매칭 결과를 Excel 파일로 출력
    시트:
    1. 기관별 라운드 배정 (매칭유형 포함) - 색상으로 매칭유형 표시
    2. 기관별 라운드 배정 - 기존 형식
    3. 라운드별 매칭
    4. 통계
    5. 미배정
    6. 경고
    """
    wb = openpyxl.Workbook()
    # 기본 시트 제거 (나중에 다시 생성할 예정)
    if len(wb.sheetnames) > 0:
        wb.remove(wb[wb.sheetnames[0]])

    # 스타일 정의
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True, size=11)
    pool1_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pool2_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    pool3_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # 기관별 라운드 배정 (매칭유형 포함) 시트 먼저 생성
    create_organization_with_type_sheet(wb, buyers, companies, matches_by_round, matched, match_type)

    # 기관별 라운드 배정 시트 생성 (기존 형식)
    create_organization_sheet(wb, buyers, companies, matches_by_round, matched, match_type)

    # ===== 시트2: 라운드별 매칭 =====
    ws1 = wb.create_sheet("라운드별 매칭")

    # 헤더
    headers = ["라운드", "바이어번호", "바이어소속", "바이어성명", "기업명", "기업주요품목", "매칭유형"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 데이터 입력
    row = 2
    for round_num in [1, 2, 3, 4]:
        for buyer_id, company_name in matches_by_round[round_num]:
            buyer = buyers[buyer_id]
            company = companies[company_name]
            pool = match_type[(buyer_id, company_name)]

            pool_map = {'pool1': '상호 관심', 'pool2': '바이어 우선', 'pool3': '기업 우선', 'pool4': '세부품목', 'pool5': '대분류'}

            ws1.cell(row=row, column=1, value=round_num)
            ws1.cell(row=row, column=2, value=buyer_id)
            ws1.cell(row=row, column=3, value=buyer['org'])
            ws1.cell(row=row, column=4, value=buyer['name'])
            ws1.cell(row=row, column=5, value=company_name)
            ws1.cell(row=row, column=6, value=company['products'])
            ws1.cell(row=row, column=7, value=pool_map[pool])

            # 색상 적용 (pool4/5는 색상 없음)
            pool_fill_map = {'pool1': pool1_fill, 'pool2': pool2_fill, 'pool3': pool3_fill}
            for col in range(1, 8):
                cell = ws1.cell(row=row, column=col)
                if pool in pool_fill_map:
                    cell.fill = pool_fill_map[pool]
                cell.border = border
                if col == 1 or col == 7:
                    cell.alignment = center_align

            row += 1

    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 25
    ws1.column_dimensions['G'].width = 12

    # ===== 시트3: 통계 =====
    ws2 = wb.create_sheet("통계")

    total_matched = len(matched)
    pool1_count = sum(1 for m in matched if match_type[m] == 'pool1')
    pool2_count = sum(1 for m in matched if match_type[m] == 'pool2')
    pool3_count = sum(1 for m in matched if match_type[m] == 'pool3')
    unmatched_count = sum(1 for buyer in buyers.values() for company in companies.values()
                         if (buyer['id'], company['name']) not in matched)

    stats = [
        ["통계 항목", "건수"],
        ["총 매칭 수", total_matched],
        ["상호희망 (Pool1)", pool1_count],
        ["바이어전용 (Pool2)", pool2_count],
        ["기업전용 (Pool3)", pool3_count],
        ["미배정 쌍", unmatched_count],
    ]

    for row_idx, row_data in enumerate(stats, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = border
            cell.alignment = center_align if col_idx == 2 else None

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12

    # ===== 시트4: 미배정 =====
    ws3 = wb.create_sheet("미배정")

    headers3 = ["바이어번호", "바이어성명", "기업명", "미배정 사유"]
    for col_idx, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row = 2
    for buyer_id, buyer in buyers.items():
        for company_name, company in companies.items():
            if (buyer_id, company_name) not in matched:
                reason = ""
                if slots[buyer_id] == 0:
                    reason = "바이어 슬롯 소진"
                else:
                    reason = "풀 미분류 또는 라운드 충돌"

                ws3.cell(row=row, column=1, value=buyer_id)
                ws3.cell(row=row, column=2, value=buyer['name'])
                ws3.cell(row=row, column=3, value=company_name)
                ws3.cell(row=row, column=4, value=reason)

                for col in range(1, 5):
                    ws3.cell(row=row, column=col).border = border

                row += 1

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 20

    # ===== 시트5: 경고 =====
    ws4 = wb.create_sheet("경고")

    headers4 = ["경고 메시지"]
    for col_idx, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row = 2
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for warning in load_warnings:
        cell = ws4.cell(row=row, column=1, value=warning)
        cell.fill = warn_fill
        cell.border = border
        row += 1

    ws4.column_dimensions['A'].width = 50

    # ===== 시트6: 특이사항 =====
    ws5 = wb.create_sheet("특이사항")

    headers5 = ["특이사항 유형", "내용"]
    for col_idx, header in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row = 2
    remark_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # 경고 메시지를 특이사항으로 추가
    for warning in load_warnings:
        ws5.cell(row=row, column=1, value="데이터 이상")
        ws5.cell(row=row, column=2, value=warning)
        for col in range(1, 3):
            ws5.cell(row=row, column=col).fill = remark_fill
            ws5.cell(row=row, column=col).border = border
        row += 1

    ws5.column_dimensions['A'].width = 15
    ws5.column_dimensions['B'].width = 60

    wb.save(output_filepath)
    print(f"[OK] 결과 저장됨: {output_filepath}")


# ==================== 메인 실행 ====================

def main():
    # 파일 경로
    buyers_file = "2026 구매담당자 리스트.xlsx"
    companies_file = "2026 참가기업 리스트.xlsx"
    output_file = "매칭결과_2026.xlsx"

    print("=" * 60)
    print("2026 공공기관 내수구매상담회 매칭 시스템")
    print("=" * 60)

    # 1. 데이터 로드
    print("\n[1] 데이터 로딩 중...")
    buyers, buyer_warnings = load_buyers(buyers_file)
    companies, company_warnings = load_companies(companies_file, buyers)
    all_warnings = buyer_warnings + company_warnings

    print(f"  [OK] 바이어 로드: {len(buyers)}명")
    print(f"  [OK] 기업 로드: {len(companies)}개사")

    # 날짜별 바이어 분포 출력
    date_distribution = {}
    for buyer_id, buyer in buyers.items():
        date = buyer.get('date', 'Unknown')
        if date not in date_distribution:
            date_distribution[date] = 0
        date_distribution[date] += 1

    print(f"\n  날짜별 바이어 분포:")
    for date in sorted(date_distribution.keys()):
        print(f"    - {date}: {date_distribution[date]}명")

    # 2. 풀 분류
    print("\n[2] 쌍 분류 중...")
    pool1, pool2, pool3, pool4, pool5 = classify_pairs(buyers, companies)
    print(f"  [OK] Pool 1 (상호 관심): {len(pool1)}쌍")
    print(f"  [OK] Pool 2 (바이어 우선): {len(pool2)}쌍")
    print(f"  [OK] Pool 3 (기업 우선): {len(pool3)}쌍")
    print(f"  [OK] Pool 4 (세부품목): {len(pool4)}쌍")
    print(f"  [OK] Pool 5 (대분류): {len(pool5)}쌍")

    # 3. 매칭 실행
    print("\n[3] 그리디 매칭 실행 중...")
    matches_by_round, matched, match_type, slots = greedy_matching(
        buyers, companies, pool1, pool2, pool3, pool4, pool5
    )

    total_matches = len(matched)
    print(f"  [OK] 총 매칭: {total_matches}쌍")
    for r in [1, 2, 3, 4]:
        print(f"    - 라운드 {r}: {len(matches_by_round[r])}쌍")

    # 4. 결과 저장
    print("\n[4] 결과 저장 중...")
    write_results(output_file, buyers, companies, matches_by_round, matched, match_type, slots, all_warnings)

    # 5. 통계
    print("\n[5] 최종 통계")
    pool1_matched = sum(1 for m in matched if match_type[m] == 'pool1')
    pool2_matched = sum(1 for m in matched if match_type[m] == 'pool2')
    pool3_matched = sum(1 for m in matched if match_type[m] == 'pool3')
    pool4_matched = sum(1 for m in matched if match_type[m] == 'pool4')
    pool5_matched = sum(1 for m in matched if match_type[m] == 'pool5')

    print(f"  [OK] Pool1 (상호 관심): {pool1_matched}쌍")
    print(f"  [OK] Pool2 (바이어 우선): {pool2_matched}쌍")
    print(f"  [OK] Pool3 (기업 우선): {pool3_matched}쌍")
    print(f"  [OK] Pool4 (세부품목): {pool4_matched}쌍")
    print(f"  [OK] Pool5 (대분류): {pool5_matched}쌍")

    if all_warnings:
        print(f"\n[WARNING] 데이터 경고: {len(all_warnings)}건")
        for warning in all_warnings:
            print(f"  - {warning}")

    print("\n" + "=" * 60)
    print("매칭 완료!")
    print("=" * 60)


if __name__ == "__main__":
    main()
